import requests
import json
import jsonpath
import openpyxl



def test_add_new_student():
    #Define the URL
    url = 'http://thetestingworldapi.com/api/studentsDetails'
    #Opn file and put it in json format
    f = open('C:/Users/liorkat/Documents/Lior Comp/API Testing/Python/json/studentMultiData.json','r')
    json_format = json.loads(f.read())

    #Open XLS file and store the values from the xls to the json file
    wb = openpyxl.load_workbook('C:/Users/liorkat/Documents/Lior Comp/API Testing/Python/json/testData.xlsx','r')
    sh = wb['Sheet1']
    rows = sh.max_row

    for i in range (2,rows + 1):
        cell_first_name = sh.cell(row = i, column = 1)
        cell_midle_name = sh.cell(row=i, column = 2)
        cell_last_name = sh.cell(row=i, column = 3)
        cell_dateOfBirth = sh.cell(row=i, column = 4)

        #update the json with the xls data
        json_format['first_name'] = cell_first_name.value
        json_format['middle_name'] = cell_midle_name.value
        json_format['last_name'] = cell_last_name.value
        json_format['date_of_birth'] = cell_dateOfBirth.value

        #post the request multipile times following the for loop
        responseRequest = requests.post(url,json_format)
        print(responseRequest.status_code)
        print(responseRequest.text)
        assert responseRequest.status_code == 201

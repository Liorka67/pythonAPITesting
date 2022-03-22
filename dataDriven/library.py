import json
import jsonpath
import requests
import openpyxl

class common:

    def __init__(self,filePath,sheetName):
        global wb
        global sh
        wb = openpyxl.load_workbook(filePath)
        sh = wb[sheetName]

    def fetch_row_counts(self):
        rows = sh.max_row
        return rows

    def fetch_column_count(self):
        col = sh.max_column
        return col

    def fetch_key_names(self):
        c= sh.max_column
        li=[]
        for i in range(1,c+1):
            cell = sh.cell(row=1,column=i)
            li.insert(i-1,cell.value)
            return li

    def update_request_with_data(self,rowNumber,jasonrequest,keyList):
        c = sh.max_column
        for i in range (1,c+1):
            cell = sh.cell(row=rowNumber,column=i)
            jasonrequest[keyList[i-1]]=cell.value
            return jasonrequest


